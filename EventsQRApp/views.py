from django.shortcuts import render
from django.template import loader
from django.http import HttpResponse
from .models import Event, Attendee, EventAttendee
from openpyxl import Workbook, load_workbook
import pandas as pd
import random, string
from django.http import JsonResponse

# Create your views here.
def ImportExcelAttendees(request):
    if request.method == "POST":
        eventObj = Event.objects.get(pk = request.POST["events"])
        df = pd.read_excel(request.FILES["file"])
        idsArr = []
        for val in df.values.tolist():
            attendee = Attendee(FirstName=val[0],MiddleName=val[1],LastName=val[2])
            attendee.save()
            idsArr.append(attendee)
       
        for val in idsArr:
             qr = ''.join(random.choices(string.ascii_letters + string.digits, k=16))
             event_attendee = EventAttendee(event=eventObj, attendee=val,qrcode=qr)
             event_attendee.save()

    
    events = Event.objects.all()
    data = {
        "events":events,
    }
    template = loader.get_template("EventsQRApp/import-attendees.html")
    return HttpResponse(template.render(data, request))

    # wb = Workbook()
    # ws = wb.active
    # ws['A1'] = 42
    # ws.append([1, 2, 3])
    # import datetime
    # ws['A2'] = datetime.datetime.now()
    # wb.save("sample.xlsx")
    # events = Event.objects.all()
    # data = {
    #     "events":events,
    # }
    # template = loader.get_template("EventsQRApp/import-attendees.html")
    # return HttpResponse(template.render(data, request))
def EventAttendeeQRCode(request,event_id):
    template = loader.get_template("EventsQRApp/event-attendee-qr.html")
    event = Event.objects.get(pk=event_id)
    event_attendees = EventAttendee.objects.filter(event=event)    
    data = {
        "event_attendees": event_attendees
    }
    return HttpResponse(template.render(data, request))

def ScanAttendeesQRCode(request,event_id,day_id):
    if request.method == "POST":
        return JsonResponse({'foo':request.POST["qrcode"]})
    event_attendees = EventAttendee.objects.filter(event_id=event_id)
    template = loader.get_template("EventsQRApp/scan-attendee-qr.html")
    event = Event.objects.get(pk=event_id)
    data = {
        "event": event,
        "event_attendees":event_attendees,
        "day": day_id
    }
    return HttpResponse(template.render(data, request))

def GenerateQRCodeAttendee(request,event_id):
    event = Event.objects.get(pk=event_id)
    attendees = Attendee.objects.exclude(
        eventattendee__in = EventAttendee.objects.filter(
            event = event
        )
    )
    template = loader.get_template("EventsQRApp/generate-qr-code-attendee.html")

    if request.method == "POST":
        qr = ''.join(random.choices(string.ascii_letters + string.digits, k=16))
        attendee = Attendee.objects.get(pk=request.POST["attendee_id"])
        event_attendee = EventAttendee(event=event,attendee=attendee,qrcode=qr)
        event_attendee.save()
        data = {
            "event": event,
            "attendees":attendees,
            "attendee":{
                "full_name": attendee.FirstName+" "+attendee.MiddleName+" "+attendee.LastName,
                "qr_code": event_attendee.qrcode
            }
        }
        return HttpResponse(template.render(data, request))

    print(event.EventName)
    data = {
        "event": event,
        "attendees":attendees
    }
    return HttpResponse(template.render(data, request))

def DisplayGeneratedQRCode(request,event_id,attendee_id):
    template = loader.get_template("EventsQRApp/display-generated-qr.html")
    event_attendee = EventAttendee.objects.get(event_id=event_id,attendee_id = attendee_id)
    data = {
        "event_attendee": event_attendee
    }
    return HttpResponse(template.render(data, request))